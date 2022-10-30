# When I wrote this code
# only I and God knew how it worked
# Now only God knows, sorry:(

# If you are debugging this code
# please increse this counter as a warning for the next person:

# totalHoursWasted = 3
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support import expected_conditions
from selenium.webdriver.support.ui import WebDriverWait
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import requests
import json
import os.path
from os import path
from datetime import datetime
import pandas
from openpyxl import load_workbook
import time
import threading

#listOfIds = [4368946, 2371, 125947, 137473, 6103603, 10581337, 6967276, 32, 3612, 3228, 2593, 4134]  # You can find new axies here: https://marketplace.axieinfinity.com/axie The id is the number after hashtag
tableAxieTransactions = jsonData = {}
l01 = l02 = l03 = l04 = l05 = l06 = l07 = l08 = l09 = l10 = l11 = l12 = l13 = l14 = l15 = l16 = l17 = l18 = l19 = l20 = l21 = l22 = l23 = l24 = l25 = l26 = l27 = ['']
fileName = 'Axies.xlsx'
auction = name = priceUSD = rarity = priceLast = priceQuick = priceReasonable = pricePatient = similarAxies = ''
url = 'https://graphql-gateway.axieinfinity.com/graphql'

priceLink = 'https://www.axie.tech/axie-pricing/'
caps = DesiredCapabilities().CHROME
caps["pageLoadStrategy"] = "normal"
options = webdriver.ChromeOptions()
options.add_argument("headless")
driver = webdriver.Chrome(desired_capabilities=caps, service=Service('C:\webdrivers\chromedriver.exe'), options=options)

def getDetails(x1, x2):  # This gets details about Axie and returns them
    return requests.post(x1, json=x2)


class AxieDetails:
    def something(self):
        self.createVariables(AxieId)  # Creates variables for API
        #print(1)
        self.getAxieDetails()  # This is used to get some Axie details and store them in tableAxieDetails
        #print(2)
        self.printSomeDetails()  # This prints the details obtained in getAxieDetails
        #print(3)
        if path.exists(fileName) != True:  # If the xlsx table doesn't exist yet, it creates it
            self.createTable()
        self.getPrices(AxieId)  # This gets the price predictions and stores them separately
        #print(4)
        self.defineTable()  # This defines the table to multiple variables (Needs everything before to finish)
        #print(5)
        self.appendToTable()  # And finally this takes variables from defineTable and puts them in xlsx table
        #print(6)

    @staticmethod
    def createVariables(Axie):
        global jsonData
        jsonData = {
            "operation": "GetAxieDetail",
            "variables": {
                "axieId": str(Axie)
            },
            "query": "query GetAxieDetail($axieId: ID!) {\n  axie(axieId: $axieId) {\n    ...AxieDetail\n    __typename\n  }\n}\n\nfragment AxieDetail on Axie {\n  id\n  image\n  class\n  chain\n  name\n  genes\n  owner\n  birthDate\n  bodyShape\n  class\n  sireId\n  sireClass\n  matronId\n  matronClass\n  stage\n  title\n  breedCount\n  level\n  figure {\n    atlas\n    model\n    image\n    __typename\n  }\n  parts {\n    ...AxiePart\n    __typename\n  }\n  stats {\n    ...AxieStats\n    __typename\n  }\n  auction {\n    ...AxieAuction\n    __typename\n  }\n  ownerProfile {\n    name\n    __typename\n  }\n  battleInfo {\n    ...AxieBattleInfo\n    __typename\n  }\n  children {\n    id\n    name\n    class\n    image\n    title\n    stage\n    __typename\n  }\n  __typename\n}\n\nfragment AxieBattleInfo on AxieBattleInfo {\n  banned\n  banUntil\n  level\n  __typename\n}\n\nfragment AxiePart on AxiePart {\n  id\n  name\n  class\n  type\n  specialGenes\n  stage\n  abilities {\n    ...AxieCardAbility\n    __typename\n  }\n  __typename\n}\n\nfragment AxieCardAbility on AxieCardAbility {\n  id\n  name\n  attack\n  defense\n  energy\n  description\n  backgroundUrl\n  effectIconUrl\n  __typename\n}\n\nfragment AxieStats on AxieStats {\n  hp\n  speed\n  skill\n  morale\n  __typename\n}\n\nfragment AxieAuction on Auction {\n  startingPrice\n  endingPrice\n  startingTimestamp\n  endingTimestamp\n  duration\n  timeLeft\n  currentPrice\n  currentPriceUSD\n  suggestedPrice\n  seller\n  listingIndex\n  state\n  __typename\n}\n"
        }
        #print(jsonData)

    @staticmethod
    def getAxieDetails():
        global tableAxieTransactions, auction, name, priceUSD
        tableAxieDetails = getDetails(url, jsonData)
        auction = tableAxieDetails.json()['data']['axie']['auction']
        name = tableAxieDetails.json()['data']['axie']['name']
        if str(auction) != 'None':
            priceUSD = tableAxieDetails.json()['data']['axie']['auction']['currentPriceUSD']
        else:
            priceUSD = 'This Axie ist not for sale'

    @staticmethod
    def printSomeDetails():

        #print(str(x.json()))

        #print('name: ' + name)
        #print('id: ' + str(AxieId))
        #if str(auction) != 'None':
            #print('price: ' + str(priceUSD) + '$')
        #else:
            #print(priceUSD)
        pass

    @staticmethod
    def createTable():
        df = pandas.DataFrame(
            {'dateTime': l01, 'name': l02, 'id': l03, 'price': l04, 'level': l05, 'hp': l06, 'skill': l07, 'speed': l08,
             'morale': l09, 'breedCount': l10, 'stage': l11, 'title': l12, 'bodyShape': l13, 'birthDate': l14,
             'class': l15, 'Eyes': l16, 'Ears': l17, 'Back': l18, 'Mouth': l19, 'Horn': l20, 'Tail': l21, 'rarity': l22,
             'priceLast': l23, 'priceQuick': l24, 'priceReasonable': l25, 'pricePatient': l26, 'similarAxies': l27})
        df.to_excel(fileName, sheet_name='sheet1', index=False)


    @staticmethod
    def getPrices(Axie):
        global rarity, priceLast, priceQuick, priceReasonable, pricePatient, similarAxies
        priceAxieLink = priceLink + str(Axie)
        # driver = webdriver.Chrome(desired_capabilities=caps, executable_path=r'C:\webdrivers\chromedriver.exe', options=options)
        # ser =


        driver.get(priceAxieLink)
        #print(3.1)
        #x = 'Calculating...'
        #while (x == 'Calculating...'):
        #    try:
        #        x = str(driver.find_element(By.XPATH, '//*[@id="root"]/div[1]/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div/div/div[1]/span').text)
        #        # print(x)
        #    except:
        #        # print('lol')
        #        pass
        #print('x: '+x)

        time.sleep(2)
        #print(3.2)


        try:
            priceLast = str(driver.find_element(By.XPATH, '//*[@id="root"]/div[1]/div[2]/div[2]/div/div/div[1]/div/div[1]/div[4]/div/div[1]/div/span').text)
        except:
            priceLast = 'N/A'

        try:
            priceQuick = str(driver.find_element(By.XPATH, '//*[@id="root"]/div[1]/div[2]/div[2]/div/div/div[1]/div/div[1]/div[4]/div/div[2]/div/span').text)
        except:
            priceQuick = 'N/A'

        try:
            priceReasonable = str(driver.find_element(By.XPATH, '//*[@id="root"]/div[1]/div[2]/div[2]/div/div/div[1]/div/div[1]/div[4]/div/div[3]/div/span').text)
        except:
            priceReasonable = 'N/A'

        try:
            pricePatient = str(driver.find_element(By.XPATH, '//*[@id="root"]/div[1]/div[2]/div[2]/div/div/div[1]/div/div[1]/div[4]/div/div[6]/div/span').text)
        except:
            pricePatient = 'N/A'

        try:
            similarAxies = str(driver.find_element(By.XPATH, '//*[@id="root"]/div[1]/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div/div/div[2]/span/span').text)
        except:
            similarAxies = 'N/A'

        try:
            rarity = str(driver.find_element(By.XPATH, '//*[@id="root"]/div[1]/div[2]/div[2]/div/div/div[1]/div/div[2]/div[1]/div/div/div[1]/span').text)
        except:
            rarity = 'N/A'


        #print(3.3)
        if priceLast != 'N/A':
            priceLast = priceLast.replace('Ξ', '')
            priceLast = str(priceLast + ' WETH')
        if priceQuick != 'N/A':
            priceQuick = priceQuick.replace('Ξ', '')
            priceQuick = str(priceQuick + ' WETH')
        if priceReasonable != 'N/A':
            priceReasonable = priceReasonable.replace('Ξ', '')
            priceReasonable = str(priceReasonable + ' WETH')
        if pricePatient != 'N/A':
            pricePatient = pricePatient.replace('Ξ', '')
            pricePatient = str(pricePatient + ' WETH')

        #print('rarity: ' + rarity)
        #print('priceLast: ' + priceLast)
        #print('priceQuick: ' + priceQuick)
        #print('priceReasonable: ' + priceReasonable)
        #print('pricePatient: ' + pricePatient)
        #print('similarAxies: ' + similarAxies)



    @staticmethod
    def defineTable():
        global l01, l02, l03, l04, l05, l06, l07, l08, l09, l10, l11, l12, l13, l14, l15, l16, l17, l18, l19, l20, l21, l22, l23, l24, l25, l26, l27
        now = datetime.now()
        l01 = [now.strftime("%d/%m/%Y %H:%M:%S")]  # dateTime
        l02 = [tableAxieTransactions.json()['data']['axie']['name']]  # name
        l03 = [str(AxieId)]  # id
        l04 = [str(priceUSD)]  # price
        l05 = [str(tableAxieTransactions.json()['data']['axie']['level'])]  # level
        l06 = [str(tableAxieTransactions.json()['data']['axie']['stats']['hp'])]  # hp
        l07 = [str(tableAxieTransactions.json()['data']['axie']['stats']['skill'])]  # skill
        l08 = [str(tableAxieTransactions.json()['data']['axie']['stats']['speed'])]  # speed
        l09 = [str(tableAxieTransactions.json()['data']['axie']['stats']['morale'])]  # morale
        l10 = [str(tableAxieTransactions.json()['data']['axie']['breedCount'])]  # breedCount
        l11 = [str(tableAxieTransactions.json()['data']['axie']['stage'])]  # stage
        l12 = [str(tableAxieTransactions.json()['data']['axie']['title'])]  # title
        l13 = [str(tableAxieTransactions.json()['data']['axie']['bodyShape'])]  # bodyShape
        l14 = [str(datetime.utcfromtimestamp(tableAxieTransactions.json()['data']['axie']['birthDate']).strftime('%Y-%m-%d %H:%M:%S'))]  # birthDate
        l15 = [str(tableAxieTransactions.json()['data']['axie']['class'])]  # class
        l16 = []  # Eyes
        l17 = []  # Ears
        l18 = []  # Back
        l19 = []  # Mouth
        l20 = []  # Horn
        l21 = []  # Tail
        l22 = [str(rarity)]  # rarity
        l23 = [str(priceLast)]  # priceLast
        l24 = [str(priceQuick)]  # priceQuick
        l25 = [str(priceReasonable)]  # priceReasonable
        l26 = [str(pricePatient)]  # pricePatient
        l27 = [str(similarAxies)]  # similarAxies

        if tableAxieTransactions.json()['data']['axie']['parts'] != []:
            for i in range(len(tableAxieTransactions.json()['data']['axie']['parts'])):
                try:
                    if "id" in tableAxieTransactions.json()['data']['axie']['parts'][i]['abilities'][0]:
                        if i == 0:
                            l16 = [str(tableAxieTransactions.json()['data']['axie']['parts'][i]['id']) + ' : ' + str(tableAxieTransactions.json()['data']['axie']['parts'][i]['abilities'][0]['id'])]
                        if i == 1:
                            l17 = [str(tableAxieTransactions.json()['data']['axie']['parts'][i]['id']) + ' : ' + str(tableAxieTransactions.json()['data']['axie']['parts'][i]['abilities'][0]['id'])]
                        if i == 2:
                            l18 = [str(tableAxieTransactions.json()['data']['axie']['parts'][i]['id']) + ' : ' + str(tableAxieTransactions.json()['data']['axie']['parts'][i]['abilities'][0]['id'])]
                        if i == 3:
                            l19 = [str(tableAxieTransactions.json()['data']['axie']['parts'][i]['id']) + ' : ' + str(tableAxieTransactions.json()['data']['axie']['parts'][i]['abilities'][0]['id'])]
                        if i == 4:
                            l20 = [str(tableAxieTransactions.json()['data']['axie']['parts'][i]['id']) + ' : ' + str(tableAxieTransactions.json()['data']['axie']['parts'][i]['abilities'][0]['id'])]
                        if i == 5:
                            l21 = [str(tableAxieTransactions.json()['data']['axie']['parts'][i]['id']) + ' : ' + str(tableAxieTransactions.json()['data']['axie']['parts'][i]['abilities'][0]['id'])]

                except:
                    if i == 0:
                        l16 = [str(tableAxieTransactions.json()['data']['axie']['parts'][i]['id'])]
                    if i == 1:
                        l17 = [str(tableAxieTransactions.json()['data']['axie']['parts'][i]['id'])]
                    if i == 2:
                        l18 = [str(tableAxieTransactions.json()['data']['axie']['parts'][i]['id'])]
                    if i == 3:
                        l19 = [str(tableAxieTransactions.json()['data']['axie']['parts'][i]['id'])]
                    if i == 4:
                        l20 = [str(tableAxieTransactions.json()['data']['axie']['parts'][i]['id'])]
                    if i == 5:
                        l21 = [str(tableAxieTransactions.json()['data']['axie']['parts'][i]['id'])]

        else:
            l16 = ['This Axie dosent have parts']
            l17 = ['This Axie dosent have parts']
            l18 = ['This Axie dosent have parts']
            l19 = ['This Axie dosent have parts']
            l20 = ['This Axie dosent have parts']
            l21 = ['This Axie dosent have parts']
        # fix title
        if l12 == ['']:
            l12 = ['None']

    @staticmethod
    def appendToTable():
        #print(l01, l02, l03, l04, l05, l06, l07, l08, l09, l10, l11, l12, l13, l14, l15, l16, l17, l18, l19, l20, l21, l22, l23, l24, l25, l26, l27)
        #print(3.401)
        df = pandas.DataFrame({'dateTime': l01, 'name': l02, 'id': l03, 'price': l04, 'level': l05, 'hp': l06, 'skill': l07, 'speed': l08, 'morale': l09, 'breedCount': l10, 'stage': l11, 'title': l12, 'bodyShape': l13, 'birthDate': l14, 'class': l15, 'Eyes': l16, 'Ears': l17, 'Back': l18, 'Mouth': l19, 'Horn': l20, 'Tail': l21, 'rarity': l22, 'priceLast': l23, 'priceQuick': l24, 'priceReasonable': l25, 'pricePatient': l26, 'similarAxies': l27})
        #print(3.41)
        book = load_workbook(fileName)
        #print(3.42)
        writer = pandas.ExcelWriter(fileName, engine='openpyxl')
        #print(3.43)
        writer.book = book
        #print(3.44)
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        #print(3.45)
        for sheetname in writer.sheets:
            df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index=False, header=False)

        writer.save()


#if __name__ == "__main__":
#    for p in range(len(listOfIds)):
#        AxieId = listOfIds[p]
#        AxieDetails()
#    print(f'\033[92mDone!!!\033[0m')

if path.exists(fileName) != True:
        AxieDetails.createTable()

def smartAxie():  # This uns AxieDetails using threading
    t1 = threading.Thread(target=AxieDetails.createVariables, args=(AxieId,))  # Creates variables for API
    t2 = threading.Thread(target=AxieDetails.getAxieDetails, args=())  # This is used to get some Axie details and store them in tableAxieDetails
    t3 = threading.Thread(target=AxieDetails.printSomeDetails, args=())  # This prints the details obtained in getAxieDetails
    t4 = threading.Thread(target=AxieDetails.getPrices, args=(AxieId,))  # This gets the price predictions and stores them separately
    #t5 = threading.Thread(target=defineTable, args=())  # This defines the table to multiple variables (Needs everything before to finish)
    #t6 = threading.Thread(target=appendToTable, args=())  # And finally this takes variables from defineTable and puts them in xlsx table
    t1.start()
    t2.start()
    t3.start()
    t4.start()
    t1.join()
    t2.join()
    t3.join()
    t4.join()
    AxieDetails.defineTable()
    AxieDetails.appendToTable()

howManyAxies = 10000
if __name__ == "__main__":
    NewId = 5429
    for p in range(howManyAxies):
        print(NewId)
        AxieId = NewId
        success = None
        while success == None:
            try:
                smartAxie()
                success = 'Done'
            except Exception as e:
                print(e)
                print('lol')
                time.sleep(10)
                pass

        NewId += 1
    print(f'\033[92mDone!!!\033[0m')
