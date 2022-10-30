import pandas
from openpyxl import load_workbook
import os.path
from os import path


fileName = 'test1.xlsx'
if path.exists(fileName) != True:
    l01 = [1, 2, 3, 4]
    l02 = [1, 2, 3, 4]
    df = pandas.DataFrame({'Hello': l01, 'hi': l02})
    df.to_excel(fileName, sheet_name='sheet1', index=False)
l01 = [5, 6, 7, 8, 9, 10]
l02 = [5, 6, 7, 8, 9, 10]
df1 = pandas.DataFrame({'smth': l01, '': l02})
book = load_workbook(fileName)
writer = pandas.ExcelWriter(fileName, engine='openpyxl')
writer.book = book
writer.sheets = {ws.title: ws for ws in book.worksheets}

for sheetname in writer.sheets:
    df1.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index=False, header=False)

writer.save()